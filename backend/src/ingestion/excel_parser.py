"""Parse Excel (.xlsx) and CSV files into structured row dictionaries."""

from __future__ import annotations

import csv
from pathlib import Path
from typing import TYPE_CHECKING, Any

import structlog
from openpyxl import load_workbook

from backend.src.models.tables import BusinessMetric

if TYPE_CHECKING:
    from collections.abc import Callable

    from sqlalchemy.orm import Session

logger = structlog.get_logger(__name__)

ParsedRow = dict[str, Any]


def parse_excel_file(filepath: str | Path) -> list[ParsedRow]:
    """Parse an Excel .xlsx file into a list of dicts keyed by header row."""
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(path)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[ParsedRow] = []
    if ws is None:
        wb.close()
        return rows

    row_iter = ws.iter_rows()
    header_row = next(row_iter, None)
    if header_row is None:
        wb.close()
        return rows

    headers = [str(cell.value) for cell in header_row]

    for row in row_iter:
        values = [cell.value for cell in row]
        if all(v is None for v in values):
            continue
        rows.append(dict(zip(headers, values, strict=False)))

    wb.close()
    logger.info("excel_parsed", filepath=str(path), row_count=len(rows))
    return rows


def parse_csv_file(filepath: str | Path) -> list[ParsedRow]:
    """Parse a CSV file into a list of dicts keyed by header row."""
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(path)

    with path.open(encoding="utf-8", newline="") as f:
        rows = [dict(row) for row in csv.DictReader(f)]

    logger.info("csv_parsed", filepath=str(path), row_count=len(rows))
    return rows


def ingest_excel_to_business_metrics(
    filepath: str | Path,
    session: Session,
    collection_id: str,
) -> int:
    """Parse an Excel file and insert rows as BusinessMetric records."""
    rows = parse_excel_file(filepath)
    if not rows:
        return 0

    for row in rows:
        metric = BusinessMetric(
            collection_id=collection_id,
            metric_name=str(row.get("metric_name", "")),
            metric_value=float(row.get("metric_value", 0)),
            unit=row.get("unit"),
            period=row.get("period"),
            category=row.get("category"),
            source_file=str(filepath),
        )
        session.add(metric)

    session.commit()
    logger.info("excel_metrics_ingested", filepath=str(filepath), count=len(rows))
    return len(rows)


PARSER_REGISTRY: dict[str, Callable[..., list[ParsedRow]]] = {
    ".xlsx": parse_excel_file,
    ".csv": parse_csv_file,
}
