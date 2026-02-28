"""Tests for Excel/CSV parser."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import pytest

from backend.src.ingestion.excel_parser import (
    PARSER_REGISTRY,
    ingest_excel_to_business_metrics,
    parse_csv_file,
    parse_excel_file,
)

FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"


class TestParseExcelFile:
    def test_parses_xlsx_rows(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["metric_name", "metric_value", "unit", "period"])
        ws.append(["Revenue", 1500000, "USD", "2025-Q4"])
        ws.append(["Headcount", 250, "people", "2025-Q4"])
        filepath = tmp_path / "kpis.xlsx"
        wb.save(filepath)

        rows = parse_excel_file(filepath)
        assert len(rows) == 2
        assert rows[0]["metric_name"] == "Revenue"
        assert rows[0]["metric_value"] == 1500000
        assert rows[1]["metric_name"] == "Headcount"

    def test_returns_list_of_parsed_row_dicts(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "value"])
        ws.append(["test", 42])
        filepath = tmp_path / "simple.xlsx"
        wb.save(filepath)

        rows = parse_excel_file(filepath)
        assert isinstance(rows, list)
        assert isinstance(rows[0], dict)
        assert "name" in rows[0]

    def test_empty_spreadsheet_returns_empty_list(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["col_a", "col_b"])
        filepath = tmp_path / "empty.xlsx"
        wb.save(filepath)

        rows = parse_excel_file(filepath)
        assert rows == []

    def test_skips_rows_with_all_none_values(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["name", "value"])
        ws.append([None, None])
        ws.append(["real", 99])
        filepath = tmp_path / "sparse.xlsx"
        wb.save(filepath)

        rows = parse_excel_file(filepath)
        assert len(rows) == 1
        assert rows[0]["name"] == "real"

    def test_raises_file_not_found(self) -> None:
        with pytest.raises(FileNotFoundError):
            parse_excel_file(Path("/nonexistent/file.xlsx"))


class TestParseCsvFile:
    def test_parses_csv_rows(self, tmp_path: Path) -> None:
        csv_content = "metric_name,metric_value,unit\nRevenue,1500000,USD\nHeadcount,250,people\n"
        filepath = tmp_path / "data.csv"
        filepath.write_text(csv_content, encoding="utf-8")

        rows = parse_csv_file(filepath)
        assert len(rows) == 2
        assert rows[0]["metric_name"] == "Revenue"
        assert rows[0]["metric_value"] == "1500000"

    def test_empty_csv_returns_empty_list(self, tmp_path: Path) -> None:
        filepath = tmp_path / "empty.csv"
        filepath.write_text("col_a,col_b\n", encoding="utf-8")

        rows = parse_csv_file(filepath)
        assert rows == []

    def test_csv_raises_file_not_found(self) -> None:
        with pytest.raises(FileNotFoundError):
            parse_csv_file(Path("/nonexistent/file.csv"))

    def test_csv_handles_quoted_fields(self, tmp_path: Path) -> None:
        csv_content = 'name,description\n"Widget A","A nice, shiny widget"\n'
        filepath = tmp_path / "quoted.csv"
        filepath.write_text(csv_content, encoding="utf-8")

        rows = parse_csv_file(filepath)
        assert len(rows) == 1
        assert rows[0]["description"] == "A nice, shiny widget"


class TestIngestExcelToBusinessMetrics:
    def _create_excel(self, tmp_path: Path) -> Path:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["metric_name", "metric_value", "unit", "period", "category"])
        ws.append(["Revenue", 1500000, "USD", "2025-Q4", "Financial"])
        ws.append(["Headcount", 250, "people", "2025-Q4", "People"])
        filepath = tmp_path / "kpis.xlsx"
        wb.save(filepath)
        return filepath

    def test_creates_business_metric_rows(self, tmp_path: Path) -> None:
        filepath = self._create_excel(tmp_path)
        mock_session = MagicMock()
        collection_id = "test-collection-id"

        count = ingest_excel_to_business_metrics(filepath, mock_session, collection_id)

        assert count == 2
        assert mock_session.add.call_count == 2
        mock_session.commit.assert_called_once()

    def test_sets_collection_id_on_each_metric(self, tmp_path: Path) -> None:
        filepath = self._create_excel(tmp_path)
        mock_session = MagicMock()
        collection_id = "coll-123"

        ingest_excel_to_business_metrics(filepath, mock_session, collection_id)

        for call in mock_session.add.call_args_list:
            metric = call[0][0]
            assert metric.collection_id == collection_id

    def test_sets_source_file_on_each_metric(self, tmp_path: Path) -> None:
        filepath = self._create_excel(tmp_path)
        mock_session = MagicMock()

        ingest_excel_to_business_metrics(filepath, mock_session, "coll")

        for call in mock_session.add.call_args_list:
            metric = call[0][0]
            assert metric.source_file == str(filepath)

    def test_returns_zero_for_empty_spreadsheet(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["metric_name", "metric_value", "unit", "period", "category"])
        filepath = tmp_path / "empty.xlsx"
        wb.save(filepath)
        mock_session = MagicMock()

        count = ingest_excel_to_business_metrics(filepath, mock_session, "coll")

        assert count == 0
        mock_session.add.assert_not_called()

    def test_coerces_metric_value_to_float(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["metric_name", "metric_value", "unit", "period", "category"])
        ws.append(["Revenue", "1500000", "USD", "2025-Q4", "Financial"])
        filepath = tmp_path / "string_val.xlsx"
        wb.save(filepath)
        mock_session = MagicMock()

        ingest_excel_to_business_metrics(filepath, mock_session, "coll")

        metric = mock_session.add.call_args[0][0]
        assert isinstance(metric.metric_value, float)
        assert metric.metric_value == 1500000.0


class TestParserRegistry:
    def test_registry_contains_xlsx(self) -> None:
        assert ".xlsx" in PARSER_REGISTRY

    def test_registry_contains_csv(self) -> None:
        assert ".csv" in PARSER_REGISTRY

    def test_xlsx_maps_to_parse_excel_file(self) -> None:
        assert PARSER_REGISTRY[".xlsx"] is parse_excel_file

    def test_csv_maps_to_parse_csv_file(self) -> None:
        assert PARSER_REGISTRY[".csv"] is parse_csv_file
